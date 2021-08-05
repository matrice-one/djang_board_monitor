#!/usr/bin/env python
# coding: utf-8

# <h1> Automating EGID data retrieval </h1>
import pandas as pd
from bs4 import BeautifulSoup
import requests
from selenium import webdriver
import openpyxl
from tkinter import *
from tkinter.ttk import *
import tkinter as tk
from tkinter import filedialog
import pickle
import os

### Code starts
root = Tk()
root.title("Daretau")
root.geometry("700x500")
root.iconbitmap('/Users/neigelinerivollat/Daretau/see_logo.ico')

path_code = os.getcwd()
## INPUTS
myLabel1 = Label(root, text = "What is the client name of your lead?").pack()
ent1 = tk.Entry(root, textvariable = myLabel1)
ent1.pack()
ent1.focus_set()

def go_to_next_entry1(event):
    ent2.focus_set() # focus_set() switches the focus to the new widget

ent1.bind('<Return>', go_to_next_entry1)

myLabel2 = Label(root, text = "Client's number?").pack()
ent2 = tk.Entry(root, textvariable = myLabel2)
ent2.pack()

def go_to_next_entry2(event):
    ent3.focus_set() # focus_set() switches the focus to the new widget

ent2.bind('<Return>', go_to_next_entry2)

myLabel3 = Label(root, text = "Client's email address?").pack()
ent3 = tk.Entry(root, textvariable = myLabel3)
ent3.pack()

def go_to_next_entry3(event):
    ent4.focus_set() # focus_set() switches the focus to the new widget

ent3.bind('<Return>', go_to_next_entry3)

myLabel4= Label(root, text = "Client's geographical address? Please make sure to follow the format: 'Street + Street number + City name. Do not forget to capitalise names.").pack()
ent4 = tk.Entry(root, textvariable = myLabel4)
ent4.pack()


try:
    with open(filename, 'rb') as fi:
        word_list = pickle.load(fi)
except:
    pass


def new_directory():
    global file_path
    global path_code
    file_path = filedialog.askopenfilename()
    filename = 'mypickle.pk'
    with open(path_code + filename, 'wb') as fi:
        # dump your data into the file
        pickle.dump(file_path, fi)


    win = tk.Toplevel()
    win.wm_title("New excel file")
    myLabel5 = Label(win,
                     text="What is the name of the sheet you wish to get the data in?").pack()
    ent5 = tk.Entry(win, textvariable=myLabel5)

    def process(event=None):
        global sheet_name
        sheet_name = ent5.get()  # get the contents of the entry widget
        win.destroy()  # for example

    ent5.bind('<Return>', process)
    ent5.pack()
    ent5.focus_set()



Button(root, text='New document directory', command=new_directory).pack(side = BOTTOM, pady = 15)

## THE PROCESS
def get_results():
    ## BASE URL
    global file_path
    url1 = 'https://map.geo.admin.ch/?lang=fr&topic=ech&bgLayer=ch.swisstopo.pixelkarte-farbe&layers=ch.swisstopo.zeitreihen,ch.bfs.gebaeude_wohnungs_register,ch.bav.haltestellen-oev,ch.swisstopo.swisstlm3d-wanderwege,ch.astra.wanderland-sperrungen_umleitungen&layers_opacity=1,1,1,0.8,0.8&layers_visibility=false,true,false,false,false&layers_timestamp=18641231,,,,&E=2499579.99&N=1118709.68&zoom=8.35204548821061'
    client_name = ent1.get()
    client_number = ent2.get()
    client_email = ent3.get()
    address = ent4.get()

    address = ent4.get()
    # Create a new instance of the Safari driver & open it
    driver = webdriver.Chrome()
    driver.get(url1)

    ## Make the drop down meny appear
    inputElement = driver.find_element_by_xpath("//*[@id='search-container']/div/form/span[2]/input")

    # Type in the search
    inputElement.send_keys(address)
    driver.implicitly_wait(3)

    ## This is to click on the first result of the second section
    #### result = driver.find_element_by_xpath("//*[@id='search-container']/div/form/span[2]/span/div[2]/div/div[2]/div[1]/div")

    ## Alternative option to avoid going on similar adress
    result = driver.find_element_by_css_selector("[title*='%s']" % address)

    ## Click on the result
    result.click()
    driver.implicitly_wait(3)

    ## Obtain the link for the new page
    elements = driver.find_element_by_xpath("//a[contains(@href, '/api.geo.admin.ch')]")
    link_to_data = elements.get_attribute("href")

    # <h2> 2- Creating a table with the data </h2>

    url2 = link_to_data

    driver.implicitly_wait(3)
    driver.get(url2)
    ## Access the url2
    r = requests.get(url2)
    bs = BeautifulSoup(r.text, 'html.parser')

    ## Target the table
    table = bs.find_all('table')[0]
    rows = table.find_all('tr')

    col = []

    ## Add all rows from the table
    for row in rows:
        csvRow = []
        for cell in row.find_all(['td', 'td']):
            csvRow.append(cell.get_text())
        col.append(csvRow)

    ## Transform this list of list into df
    df = pd.DataFrame(col).transpose()
    df.columns = df.iloc[0]
    df = df[1:]

    # We add the url, client name, number and email plus adress data.

    df['url'] = url2
    df.insert(0, 'client_name', client_name)
    df.insert(1, 'client_number', client_number)
    df.insert(2, 'client_email', client_email)

    street_name_and_number = df['Désignation de la rue FR'] + ' ' + df['N° d’entrée du bâtiment']
    df.insert(6, 'street_name_and_number', street_name_and_number)


    ## We reduce the table to only the columns of interest and save it.
    df_interest = df[['client_name', 'client_number', 'client_email', 'Abréviation du canton', 'Nom le la commune',
                      'street_name_and_number', 'NPA', "Id. fédéral de bâtiment (EGID)", "N° d'immeuble", "Type d'immeuble",
                      "Classe de bâtiment", "Année de construction du bâtiment", "Epoque de construction",
                      "Surface du bâtiment [m2]", "Nombre de niveaux", 'url']]


    new_list = df_interest.values.tolist()
    new_list = new_list[0]

    # <h2> 3- Add this data to the existing table </h2>
    path_customer = file_path


    wb = openpyxl.load_workbook(path_customer)

    ## We open the table to populate and select the sheet of interest
    wb = openpyxl.load_workbook(path_customer)
    sheet = wb[sheet_name]

    ## We need to define this to know where we will have to put the next data.
    max_row = sheet.max_row
    max_col = sheet.max_column

    ## Now we can add a row
    new_list = new_list
    i = 0
    ## Adds each value of the list to the last row + 1
    for col_num in range(1, max_col + 1):
        new_value = sheet.cell(row=max_row + 1, column=col_num)
        new_value.value = new_list[i]
        i += 1

    max_row = sheet.max_row
    max_row

    wb.save(path_customer)
    wb.close()


add = Button(root, text="Launch search", command=get_results).pack()

root.mainloop()